from flask import Flask, request, jsonify, send_from_directory
import sqlite3
import psycopg2
import psycopg2.extras
import os
from dotenv import load_dotenv
from datetime import datetime
import json

# Cargar variables de entorno
load_dotenv()

app = Flask(__name__, static_folder='public', static_url_path='')

# Detectar si estamos en Railway (PostgreSQL) o local (SQLite)
DATABASE_URL = os.getenv('DATABASE_URL')
USE_POSTGRES = DATABASE_URL is not None

# Debug: Mostrar qué base de datos se está usando
print(f"🔍 DATABASE_URL encontrada: {DATABASE_URL is not None}")
print(f"🔍 Usando {'PostgreSQL' if USE_POSTGRES else 'SQLite'}")

# Ruta de la base de datos SQLite (solo para local)
DB_PATH = 'gestion_tutor.db'

def get_db_connection():
    """Crear conexión a PostgreSQL o SQLite según el entorno"""
    if USE_POSTGRES:
        # Conexión a PostgreSQL (Railway)
        conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.DictCursor)
        return conn
    else:
        # Conexión a SQLite (local)
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        return conn

def dict_from_row(row):
    """Convertir fila de DB a diccionario"""
    if USE_POSTGRES:
        return dict(row)
    else:
        return dict(row)

def get_authorized_user_ids(user_id):
    """Obtener IDs de usuarios autorizados según jerarquía"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener información del usuario
        cursor.execute(
            "SELECT id, role, supervisor_id FROM users WHERE id = %s",
            (user_id,)
        )
        user = cursor.fetchone()
        
        if not user:
            return []
        
        if user['role'] == 'admin':
            # Admin ve a sus supervisados Y a sí mismo
            cursor.execute(
                "SELECT id FROM users WHERE supervisor_id = %s",
                (user_id,)
            )
            supervised = cursor.fetchall()
            supervised_ids = [row['id'] for row in supervised]
            # Incluir al admin mismo en la lista
            supervised_ids.append(user_id)
            return supervised_ids
        else:
            # Usuario regular solo se ve a sí mismo
            return [user_id]
    finally:
        conn.close()

# ==================== RUTAS ESTÁTICAS ====================

@app.route('/')
def index():
    return send_from_directory('public', 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory('public', path)

# ==================== API: AUTH ====================

@app.route('/api/login', methods=['POST', 'OPTIONS'])
def login():
    if request.method == 'OPTIONS':
        return '', 204
    
    try:
        data = request.get_json(silent=True)
        if not data:
             return jsonify({'success': False, 'message': 'Invalid JSON body'}), 400

        username = data.get('username')
        password = data.get('password')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        try:
            cursor.execute(
                """SELECT id, username, role, branch_id 
                   FROM users 
                   WHERE username = %s AND password_hash = %s""",
                (username, password)
            )
            user = cursor.fetchone()
            
            if user:
                return jsonify({
                    'success': True,
                    'user': {
                        'id': user['id'],
                        'username': user['username'],
                        'role': user['role'],
                        'branch_id': user['branch_id']
                    }
                })
            else:
                return jsonify({'success': False, 'message': 'Credenciales inválidas'}), 401
        finally:
            conn.close()
    except Exception as e:
        print(f"Login error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

# ==================== API: DASHBOARD ====================



# ==================== API: DASHBOARD DETAILS ====================

def get_dashboard_daily_detail(status):
    user_id = request.args.get('userId')
    if not user_id:
        return jsonify({'success': False, 'message': 'userId requerido'}), 400
    
    authorized_ids = get_authorized_user_ids(int(user_id))
    if not authorized_ids:
        return jsonify({'success': True, 'data': []})
        
    placeholders = ','.join(['%s' for _ in authorized_ids])
    
    # Adjust for Mexico time (UTC-6)
    from datetime import timedelta
    today = (datetime.utcnow() - timedelta(hours=6)).date()
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(f"""
            SELECT e.id, e.full_name, b.name as branch_name, a.comment
            FROM attendance a
            JOIN employees e ON a.employee_id = e.id
            LEFT JOIN branches b ON e.branch_id = b.id
            JOIN attendance_roster ar ON e.id = ar.employee_id
            WHERE a.date = %s 
            AND a.status = %s 
            AND ar.added_by_user_id IN ({placeholders})
            ORDER BY e.full_name ASC
        """, [today, status] + authorized_ids)
        
        data = []
        for row in cursor.fetchall():
            data.append(dict_from_row(row))
            
        return jsonify({'success': True, 'data': data})
    finally:
        conn.close()

def get_dashboard_range_detail(status):
    user_id = request.args.get('userId')
    if not user_id:
        return jsonify({'success': False, 'message': 'userId requerido'}), 400
    
    authorized_ids = get_authorized_user_ids(int(user_id))
    if not authorized_ids:
        return jsonify({'success': True, 'data': []})
        
    placeholders = ','.join(['%s' for _ in authorized_ids])
    
    # Adjust for Mexico time (UTC-6)
    from datetime import timedelta
    today = (datetime.utcnow() - timedelta(hours=6)).date()
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(f"""
            SELECT DISTINCT e.id, e.full_name, b.name as branch_name, a.start_date, a.end_date, a.comment, a.permission_type
            FROM attendance a
            JOIN employees e ON a.employee_id = e.id
            LEFT JOIN branches b ON e.branch_id = b.id
            JOIN attendance_roster ar ON e.id = ar.employee_id
            WHERE a.status = %s 
            AND a.start_date <= %s 
            AND a.end_date >= %s
            AND ar.added_by_user_id IN ({placeholders})
            ORDER BY e.full_name ASC
        """, [status, today, today] + authorized_ids)
        
        data = []
        for row in cursor.fetchall():
            item = dict_from_row(row)
            # Format dates
            if item.get('start_date'):
                item['start_date'] = item['start_date'].isoformat() if hasattr(item['start_date'], 'isoformat') else str(item['start_date'])
            if item.get('end_date'):
                item['end_date'] = item['end_date'].isoformat() if hasattr(item['end_date'], 'isoformat') else str(item['end_date'])
            data.append(item)
            
        return jsonify({'success': True, 'data': data})
    finally:
        conn.close()

@app.route('/api/dashboard/stats', methods=['GET', 'OPTIONS'])
def dashboard_stats():
    if request.method == 'OPTIONS':
        return '', 204
        
    user_id = request.args.get('userId')
    if not user_id:
        return jsonify({'success': False, 'message': 'userId requerido'}), 400
    
    authorized_ids = get_authorized_user_ids(int(user_id))
    if not authorized_ids:
        return jsonify({'success': True, 'data': {
            'asistencias': 0,
            'faltas': 0,
            'vacaciones': 0,
            'permisos': 0,
            'incapacidades': 0,
            'incidencias_activas': 0
        }})
    
    placeholders = ','.join(['%s' for _ in authorized_ids])
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Get today's date adjusted for Mexico time (UTC-6)
        from datetime import timedelta
        today = (datetime.utcnow() - timedelta(hours=6)).date()
        
        # Count attendances by status for today
        cursor.execute(f"""
            SELECT status, COUNT(*) as count
            FROM attendance
            WHERE date = %s
            AND employee_id IN (
                SELECT employee_id FROM attendance_roster
                WHERE added_by_user_id IN ({placeholders})
            )
            GROUP BY status
        """, [today] + authorized_ids)
        
        stats = {
            'asistencias': 0,
            'faltas': 0,
            'vacaciones': 0,
            'permisos': 0,
            'incapacidades': 0,
            'incidencias_activas': 0
        }
        
        for row in cursor:
            status = row['status']
            count = row['count']
            
            if status == 'present':
                stats['asistencias'] = count
            elif status == 'absent':
                stats['faltas'] = count
            elif status == 'vacation':
                stats['vacaciones'] = count
            elif status == 'permission':
                stats['permisos'] = count
            elif status == 'incapacity':
                stats['incapacidades'] = count
        
        # Count active incidents (linked to employees in roster via reported_by)
        cursor.execute(f"""
            SELECT COUNT(DISTINCT i.id) as count
            FROM incidents i
            JOIN attendance_roster ar ON i.reported_by = ar.employee_id
            WHERE ar.added_by_user_id IN ({placeholders})
            AND i.status IN ('pending', 'in_progress', 'EN PROCESO')
        """, authorized_ids)
        result = cursor.fetchone()
        if result:
            stats['incidencias_activas'] = result['count']
        
        return jsonify({'success': True, 'data': stats})
    finally:
        conn.close()

@app.route('/api/dashboard/absences', methods=['GET', 'OPTIONS'])
def dashboard_absences():
    if request.method == 'OPTIONS': return '', 204
    return get_dashboard_daily_detail('absent')

@app.route('/api/dashboard/vacations', methods=['GET', 'OPTIONS'])
def dashboard_vacations():
    if request.method == 'OPTIONS': return '', 204
    return get_dashboard_range_detail('vacation')

@app.route('/api/dashboard/permissions', methods=['GET', 'OPTIONS'])
def dashboard_permissions():
    if request.method == 'OPTIONS': return '', 204
    return get_dashboard_range_detail('permission')

@app.route('/api/dashboard/sick-leaves', methods=['GET', 'OPTIONS'])
def dashboard_sick_leaves():
    if request.method == 'OPTIONS': return '', 204
    return get_dashboard_range_detail('incapacity')

@app.route('/api/dashboard/active-incidents', methods=['GET', 'OPTIONS'])
def dashboard_active_incidents():
    if request.method == 'OPTIONS': return '', 204
    
    user_id = request.args.get('userId')
    if not user_id:
        return jsonify({'success': False, 'message': 'userId requerido'}), 400
    
    authorized_ids = get_authorized_user_ids(int(user_id))
    if not authorized_ids:
        return jsonify({'success': True, 'data': []})
        
    placeholders = ','.join(['%s' for _ in authorized_ids])
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(f"""
            SELECT DISTINCT i.*, e.full_name as reported_by_name, b.name as branch_name
            FROM incidents i
            JOIN employees e ON i.reported_by = e.id
            LEFT JOIN branches b ON i.branch_id = b.id
            JOIN attendance_roster ar ON i.reported_by = ar.employee_id
            WHERE i.status IN ('pending', 'in_progress', 'EN PROCESO')
            AND ar.added_by_user_id IN ({placeholders})
            ORDER BY i.created_at DESC
        """, authorized_ids)
        
        data = []
        for row in cursor.fetchall():
            item = dict(row)
            for k, v in item.items():
                if hasattr(v, 'isoformat'):
                    item[k] = v.isoformat()
            data.append(item)
            
        return jsonify({'success': True, 'data': data})
    finally:
        conn.close()

@app.route('/api/incidents', methods=['GET', 'POST', 'OPTIONS'])
def incidents_list():
    if request.method == 'OPTIONS': 
        return '', 204
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        if request.method == 'GET':
            user_id = request.args.get('userId')
            if not user_id:
                return jsonify({'success': False, 'message': 'userId requerido'}), 400
                
            authorized_ids = get_authorized_user_ids(int(user_id))
            if not authorized_ids:
                return jsonify({'success': True, 'data': []})
                
            placeholders = ','.join(['%s' for _ in authorized_ids])
            
            # Filter incidents by supervised employees (via roster)
            cursor.execute(f"""
                SELECT i.*, e.full_name as reported_by_name, b.name as branch_name 
                FROM incidents i
                JOIN employees e ON i.reported_by = e.id
                LEFT JOIN branches b ON i.branch_id = b.id
                JOIN attendance_roster ar ON i.reported_by = ar.employee_id
                WHERE ar.added_by_user_id IN ({placeholders})
                ORDER BY i.created_at DESC
            """, authorized_ids)
            
            data = []
            for row in cursor.fetchall():
                item = dict(row)
                for k, v in item.items():
                    if hasattr(v, 'isoformat'):
                        item[k] = v.isoformat()
                data.append(item)
            return jsonify({'success': True, 'data': data})
        
        elif request.method == 'POST':
            data = request.json
            
            # Convertir cadenas vacías a None para campos de fecha
            start_date = data.get('start_date') or None
            end_date = data.get('end_date') or None
            branch_id = data.get('branch_id') or None
            reported_by = data.get('reported_by') or None
            
            cursor.execute("""
                INSERT INTO incidents 
                (branch_id, reported_by, type, status, description, start_date, end_date)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                branch_id,
                reported_by,
                data.get('type'),
                data.get('status', 'pending'),
                data.get('description'),
                start_date,
                end_date
            ))
            conn.commit()
            return jsonify({'success': True, 'message': 'Incidencia creada'})
    finally:
        conn.close()

@app.route('/api/incidents/<int:id>', methods=['PUT', 'DELETE', 'OPTIONS'])
def incident_detail(id):
    if request.method == 'OPTIONS':
        return '', 204
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        if request.method == 'DELETE':
            cursor.execute("DELETE FROM incidents WHERE id = %s", (id,))
            conn.commit()
            return jsonify({'success': True, 'message': 'Incidencia eliminada'})
            
        elif request.method == 'PUT':
            data = request.json
            print(f"PUT /api/incidents/{id} - Received data: {data}")
            
            # Convert empty strings to None for date fields
            start_date = data.get('start_date') or None
            end_date = data.get('end_date') or None
            
            # Frontend sends 'employee_id' but we need 'reported_by'
            # Also handle 'branch_id' which may not be sent
            reported_by = data.get('reported_by') or data.get('employee_id') or None
            branch_id = data.get('branch_id') or None
            
            print(f"Parsed values - branch_id: {branch_id}, reported_by: {reported_by}, type: {data.get('type')}, status: {data.get('status')}")
            
            cursor.execute("""
                UPDATE incidents 
                SET branch_id = %s,
                    reported_by = %s,
                    type = %s,
                    status = %s,
                    description = %s,
                    start_date = %s,
                    end_date = %s
                WHERE id = %s
            """, (
                branch_id,
                reported_by,
                data.get('type'),
                data.get('status'),
                data.get('description'),
                start_date,
                end_date,
                id
            ))
            conn.commit()
            print(f"Successfully updated incident {id}")
            return jsonify({'success': True, 'message': 'Incidencia actualizada'})
    except Exception as e:
        print(f"Error in incident_detail: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

# ==================== API: SUPERVISED TUTORS ====================

@app.route('/api/supervised-tutors', methods=['GET', 'OPTIONS'])
def get_supervised_tutors():
    if request.method == 'OPTIONS':
        return '', 204
        
    user_id = request.args.get('userId')
    if not user_id:
        return jsonify({'success': False, 'message': 'userId requerido'}), 400
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Get user role
        cursor.execute("SELECT role FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        
        if not user or user['role'] != 'admin':
            return jsonify({'success': True, 'tutors': []})
        
        # Get supervised tutors
        cursor.execute("""
            SELECT id, username, role
            FROM users
            WHERE supervisor_id = %s AND role = 'tutor_analista'
            ORDER BY username ASC
        """, (user_id,))
        
        tutors = []
        for row in cursor:
            tutors.append({
                'id': row['id'],
                'username': row['username'],
                'role': row['role']
            })
        
        return jsonify({'success': True, 'tutors': tutors})
    finally:
        conn.close()

# ==================== API: BRANCHES ====================

@app.route('/api/branches', methods=['GET', 'POST', 'OPTIONS'])
def get_branches():
    if request.method == 'OPTIONS':
        return '', 204
        
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        if request.method == 'GET':
            cursor.execute("SELECT * FROM branches ORDER BY name ASC")
            data = [dict(row) for row in cursor.fetchall()]
            return jsonify({'success': True, 'data': data})
        
        elif request.method == 'POST':
            data = request.json
            branch_name = data.get('name')
            
            if not branch_name:
                return jsonify({'success': False, 'message': 'El nombre de la sucursal es obligatorio'}), 400
            
            # Verificar si ya existe una sucursal con ese nombre
            cursor.execute("SELECT id, name FROM branches WHERE name = %s", (branch_name,))
            existing = cursor.fetchone()
            
            if existing:
                # Si ya existe, retornar el ID existente en lugar de error
                return jsonify({'success': True, 'message': 'Sucursal ya existe', 'data': {'id': existing['id'], 'name': existing['name']}})
            
            try:
                # Insertar la nueva sucursal sin especificar ID (usa DEFAULT)
                cursor.execute("INSERT INTO branches (name) VALUES (%s) RETURNING id", (branch_name,))
                new_id = cursor.fetchone()['id']
                conn.commit()
                
                return jsonify({'success': True, 'message': 'Sucursal creada correctamente', 'data': {'id': new_id, 'name': branch_name}})
            except Exception as insert_error:
                conn.rollback()
                # Si hay error de secuencia, intentar arreglarlo
                if 'duplicate key' in str(insert_error) or 'unique constraint' in str(insert_error):
                    try:
                        # Arreglar la secuencia
                        cursor.execute("SELECT setval('branches_id_seq', (SELECT MAX(id) FROM branches))")
                        conn.commit()
                        # Reintentar el insert
                        cursor.execute("INSERT INTO branches (name) VALUES (%s) RETURNING id", (branch_name,))
                        new_id = cursor.fetchone()['id']
                        conn.commit()
                        return jsonify({'success': True, 'message': 'Sucursal creada correctamente', 'data': {'id': new_id, 'name': branch_name}})
                    except Exception as retry_error:
                        conn.rollback()
                        print(f"Error al crear sucursal después de arreglar secuencia: {retry_error}")
                        return jsonify({'success': False, 'message': f'Error al crear sucursal: {str(retry_error)}'}), 500
                else:
                    print(f"Error al crear sucursal: {insert_error}")
                    return jsonify({'success': False, 'message': f'Error al crear sucursal: {str(insert_error)}'}), 500
    except Exception as e:
        print(f"Error general en get_branches: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

# ==================== API: EMPLOYEES ====================

@app.route('/api/employees', methods=['GET', 'POST', 'OPTIONS'])
def employees():
    if request.method == 'OPTIONS':
        return '', 204
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        if request.method == 'GET':
            cursor.execute("""
                SELECT e.id, e.full_name, e.branch_id, e.hire_date, b.name as branch_name, e.status
                FROM employees e
                LEFT JOIN branches b ON e.branch_id = b.id
                ORDER BY e.full_name ASC
            """)
            
            employees = []
            for row in cursor:
                employees.append({
                    'id': row['id'],
                    'full_name': row['full_name'],
                    'branch_id': row['branch_id'],
                    'hire_date': row['hire_date'].isoformat() if row['hire_date'] and hasattr(row['hire_date'], 'isoformat') else row['hire_date'],
                    'branch_name': row['branch_name'],
                    'status': row['status']
                })
            
            return jsonify({'success': True, 'data': employees})
        
        elif request.method == 'POST':
            data = request.json
            hire_date = data.get('hire_date') or None
            
            try:
                cursor.execute("""
                    INSERT INTO employees (full_name, branch_id, hire_date, status)
                    VALUES (%s, %s, %s, 'active')
                """, (data.get('full_name'), data.get('branch_id'), hire_date))
                conn.commit()
                
                return jsonify({'success': True, 'message': 'Empleado creado'})
            except Exception as insert_error:
                conn.rollback()
                # Si hay error de secuencia, intentar arreglarlo
                if 'duplicate key' in str(insert_error) or 'unique constraint' in str(insert_error):
                    try:
                        # Arreglar la secuencia
                        cursor.execute("SELECT setval('employees_id_seq', (SELECT MAX(id) FROM employees))")
                        conn.commit()
                        # Reintentar el insert
                        cursor.execute("""
                            INSERT INTO employees (full_name, branch_id, hire_date, status)
                            VALUES (%s, %s, %s, 'active')
                        """, (data.get('full_name'), data.get('branch_id'), hire_date))
                        conn.commit()
                        return jsonify({'success': True, 'message': 'Empleado creado'})
                    except Exception as retry_error:
                        conn.rollback()
                        print(f"Error al crear empleado después de arreglar secuencia: {retry_error}")
                        return jsonify({'success': False, 'message': f'Error al crear empleado: {str(retry_error)}'}), 500
                else:
                    print(f"Error al crear empleado: {insert_error}")
                    return jsonify({'success': False, 'message': f'Error al crear empleado: {str(insert_error)}'}), 500
    except Exception as e:
        print(f"Error general en employees: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/employees/<int:id>', methods=['PUT', 'DELETE', 'OPTIONS'])
def employee_detail(id):
    if request.method == 'OPTIONS':
        return '', 204
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        if request.method == 'DELETE':
            cursor.execute("DELETE FROM employees WHERE id = %s", (id,))
            conn.commit()
            return jsonify({'success': True, 'message': 'Empleado eliminado'})
            
        elif request.method == 'PUT':
            data = request.json
            hire_date = data.get('hire_date') or None
            
            cursor.execute("""
                UPDATE employees 
                SET full_name = %s,
                    branch_id = %s,
                    hire_date = %s
                WHERE id = %s
            """, (data.get('full_name'), data.get('branch_id'), hire_date, id))
            conn.commit()
            
            return jsonify({'success': True, 'message': 'Empleado actualizado'})
    except Exception as e:
        print(f"Error in employee_detail: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

# ==================== API: ATTENDANCE ====================

@app.route('/api/attendance', methods=['GET', 'POST', 'OPTIONS'])
def attendance():
    if request.method == 'OPTIONS':
        return '', 204
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        if request.method == 'GET':
            user_id = request.args.get('userId')
            
            if not user_id:
                return jsonify({'success': False, 'message': 'userId requerido'}), 400
            
            authorized_ids = get_authorized_user_ids(int(user_id))
            
            if not authorized_ids:
                return jsonify({'success': False, 'message': 'No autorizado'}), 403
            
            placeholders = ','.join(['%s' for _ in authorized_ids])
            
            # Obtener roster (empleados en la lista de asistencia)
            # Tutores ven solo lo que ellos agregan
            # Admins ven solo lo que sus tutores supervisados agregan
            cursor.execute(f"""
                SELECT DISTINCT e.id, e.full_name, e.branch_id, b.name as branch_name
                FROM attendance_roster ar
                JOIN employees e ON ar.employee_id = e.id
                LEFT JOIN branches b ON e.branch_id = b.id
                WHERE ar.added_by_user_id IN ({placeholders})
                ORDER BY e.full_name ASC
            """, authorized_ids)
            
            employees = []
            for row in cursor:
                employees.append({
                    'id': row['id'],
                    'full_name': row['full_name'],
                    'branch_id': row['branch_id'],
                    'branch_name': row['branch_name'],
                    'marks': {}  # Will be populated with attendance records
                })
            
            if not employees:
                return jsonify({'success': True, 'data': []})
            
            # Obtener todos los registros de asistencia para estos empleados
            employee_ids = [emp['id'] for emp in employees]
            emp_placeholders = ','.join(['%s' for _ in employee_ids])
            
            # Obtener registros de asistencia filtrados por mes/año si se proporcionan
            month = request.args.get('month')
            year = request.args.get('year')
            
            query = """
                SELECT a.employee_id, a.date, a.status, a.comment,
                       a.arrival_time, a.permission_type, a.start_date, a.end_date
                FROM attendance a
                WHERE a.employee_id IN ({})
            """.format(emp_placeholders)
            
            params = employee_ids.copy()
            
            if month is not None and year is not None:
                # Filtrar por mes específico
                # Nota: month en JS es 0-11, aquí ajustamos si es necesario o asumimos formato
                # Asumiendo que frontend manda 0 para Enero, en SQL EXTRACT(MONTH) es 1 para Enero
                # Ajustamos sumando 1 al mes recibido
                try:
                    target_month = int(month) + 1
                    target_year = int(year)
                    query += " AND EXTRACT(MONTH FROM a.date) = %s AND EXTRACT(YEAR FROM a.date) = %s"
                    params.extend([target_month, target_year])
                except ValueError:
                    pass # Ignorar si no son números válidos
            
            query += " ORDER BY a.date DESC"
            
            cursor.execute(query, params)
            
            # Organizar las marcas por empleado y fecha
            for row in cursor:
                emp_id = row['employee_id']
                date_str = row['date'].isoformat() if hasattr(row['date'], 'isoformat') else str(row['date'])
                
                # Encontrar el empleado correspondiente
                for emp in employees:
                    if emp['id'] == emp_id:
                        emp['marks'][date_str] = {
                            'status': row['status'],
                            'comment': row['comment'],
                            'arrival_time': row['arrival_time'],
                            'permission_type': row['permission_type'],
                            'start_date': row['start_date'].isoformat() if row['start_date'] and hasattr(row['start_date'], 'isoformat') else row['start_date'],
                            'end_date': row['end_date'].isoformat() if row['end_date'] and hasattr(row['end_date'], 'isoformat') else row['end_date']
                        }
                        break
            
            return jsonify({
                'success': True,
                'data': employees
            })
        
        elif request.method == 'POST':
            data = request.json
            cursor.execute("""
                INSERT INTO attendance 
                (employee_id, date, status, comment, arrival_time, 
                 permission_type, start_date, end_date)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                data.get('employee_id'),
                data.get('date'),
                data.get('status'),
                data.get('comment'),
                data.get('arrival_time'),
                data.get('permission_type'),
                data.get('start_date'),
                data.get('end_date')
            ))
            conn.commit()
            
            return jsonify({'success': True})
    finally:
        conn.close()

@app.route('/api/attendance/archived-months', methods=['GET', 'OPTIONS'])
def archived_months():
    if request.method == 'OPTIONS':
        return '', 204
    
    user_id = request.args.get('userId')
    if not user_id:
        return jsonify({'success': False, 'message': 'userId requerido'}), 400
        
    authorized_ids = get_authorized_user_ids(int(user_id))
    if not authorized_ids:
        return jsonify({'success': True, 'data': []})
        
    placeholders = ','.join(['%s' for _ in authorized_ids])
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(f"""
            SELECT 
                CAST(EXTRACT(MONTH FROM a.date) AS INTEGER) as month, 
                CAST(EXTRACT(YEAR FROM a.date) AS INTEGER) as year, 
                COUNT(*) as record_count
            FROM attendance a
            JOIN attendance_roster ar ON a.employee_id = ar.employee_id
            WHERE ar.added_by_user_id IN ({placeholders})
            GROUP BY year, month
            ORDER BY year DESC, month DESC
        """, authorized_ids)
        
        data = []
        for row in cursor.fetchall():
            data.append({
                'month': int(row['month']) - 1,
                'year': int(row['year']),
                'record_count': row['record_count']
            })
            
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        print(f"Error en archived_months: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/attendance/roster', methods=['POST', 'OPTIONS'])
def add_to_roster():
    if request.method == 'OPTIONS':
        return '', 204
        
    data = request.json
    employee_id = data.get('employee_id')
    user_id = data.get('userId')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            INSERT INTO attendance_roster (employee_id, added_by_user_id)
            VALUES (%s, %s)
            ON CONFLICT (employee_id) DO NOTHING
        """, (employee_id, user_id))
        conn.commit()
        
        return jsonify({'success': True})
    finally:
        conn.close()

@app.route('/api/attendance/roster/<int:employee_id>', methods=['DELETE', 'OPTIONS'])
def delete_from_roster(employee_id):
    if request.method == 'OPTIONS':
        return '', 204
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            DELETE FROM attendance_roster 
            WHERE employee_id = %s
        """, (employee_id,))
        conn.commit()
        
        return jsonify({'success': True, 'message': 'Colaborador eliminado del roster'})
    finally:
        conn.close()


def check_accumulated_delays(cursor, employee_id, date_obj):
    """Check for 3 accumulated delays in the current month and create alerts"""
    try:
        # Get start and end of the month for the given date
        # date_obj might be a string or datetime.date
        if isinstance(date_obj, str):
            try:
                date_obj = datetime.strptime(date_obj, '%Y-%m-%d').date()
            except:
                pass # Try to use it or fail gracefully
                
        if not hasattr(date_obj, 'year'):
            return

        month = date_obj.month
        year = date_obj.year
        
        # Count delays in this month
        cursor.execute("""
            SELECT date, comment 
            FROM attendance 
            WHERE employee_id = %s 
            AND status = 'delay'
            AND EXTRACT(MONTH FROM date) = %s
            AND EXTRACT(YEAR FROM date) = %s
            ORDER BY date ASC
        """, (employee_id, month, year))
        
        records = cursor.fetchall()
        count = len(records)
        
        # Trigger on 3rd delay, 6th delay, etc? 
        # User asked for "3 retardos". Let's trigger if count >= 3.
        # To avoid spam, we check if alert already exists for this specific month/milestone.
        
        if count < 3:
            return

        # Get employee info
        cursor.execute("SELECT full_name FROM employees WHERE id = %s", (employee_id,))
        emp_res = cursor.fetchone()
        if not emp_res: return
        emp_name = emp_res['full_name']
        
        # Get tutor and supervisor
        cursor.execute("""
            SELECT ar.added_by_user_id, u.supervisor_id
            FROM attendance_roster ar
            JOIN users u ON ar.added_by_user_id = u.id
            WHERE ar.employee_id = %s
        """, (employee_id,))
        
        roster_info = cursor.fetchone()
        if not roster_info: return
            
        tutor_id = roster_info['added_by_user_id']
        supervisor_id = roster_info['supervisor_id']
        
        # Prepare details
        month_names = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        month_name = month_names[month]
        
        latest_date_str = date_obj.isoformat()
        
        alert_details = json.dumps({
            'type': '3_delays', # Keep type for frontend compatibility or change to 'accumulated_delays'
            'subtype': 'accumulated',
            'employee_name': emp_name,
            'month': month_name,
            'year': year,
            'count': count,
            'latest_date': latest_date_str,
            'delays': [
                {
                    'date': r['date'].isoformat() if hasattr(r['date'], 'isoformat') else str(r['date']), 
                    'comment': r['comment']
                } 
                for r in records
            ]
        })
        
        unique_key = f"acc_delays_{year}_{month}_count_{count}" # Alert per count? Or just once per month?
        # If we alert on 3, we don't want to alert on 3 again.
        # If they delete one and add it back, maybe we do? 
        # Let's alert if we hit 3. If they hit 4, do we alert? 
        # User said "has 3 delays". Usually implies a threshold. 
        # Let's target exactly count >= 3 and alert ONCE per month for "3 delays".
        # If they reach 6, maybe another? Let's stick to >= 3 and trigger if not sent for this month.
        
        # Actually proper logic: Trigger whenever they hit a multiple of 3? 
        # Or just "has 3". Let's assume threshold of 3.
        
        def create_alert_for_user(uid):
            if not uid: return
            
            # Check if alert for this month already exists
            # We use a unique string in 'details' to identify it essentially, or check logic.
            # Let's look for alerts created this month for this employee with type 3_delays
            # AND containing the specific month/year in details to be safe.
            
            searchTerm = f'"month": "{month_name}", "year": {year}'
            
            cursor.execute("""
                SELECT id, details FROM alerts 
                WHERE user_id = %s AND employee_id = %s AND details LIKE %s
            """, (uid, employee_id, f'%{searchTerm}%'))
            
            existing = cursor.fetchall()
            
            # If we already have an alert for this month, do we send another?
            # If the user asks for "3 consecutive", and we switched to "accumulated",
            # usually you get one warning "This user reached 3 delays in Jan".
            # You don't need one for 4, 5. 
            # Unless they hit 6 (another 3). 
            # Let's keep it simple: One alert per month when they hit 3+.
            
            if existing: 
                return
                
            cursor.execute("""
                INSERT INTO alerts (user_id, employee_id, details)
                VALUES (%s, %s, %s)
            """, (uid, employee_id, alert_details))

        create_alert_for_user(tutor_id)
        if supervisor_id:
            create_alert_for_user(supervisor_id)
            
    except Exception as e:
        print(f"Error checking accumulated delays: {e}")
        import traceback
        traceback.print_exc()

@app.route('/api/attendance/mark', methods=['POST', 'OPTIONS'])

def mark_attendance():
    if request.method == 'OPTIONS':
        return '', 204
        
    data = request.json
    employee_id = data.get('employee_id')
    date = data.get('date')
    status = data.get('status')
    comment = data.get('comment', '')
    arrival_time = data.get('arrival_time', '')
    permission_type = data.get('permission_type', '')
    start_date = data.get('start_date') or None
    end_date = data.get('end_date') or None
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        if status == 'none':
            # Delete the attendance record
            cursor.execute("""
                DELETE FROM attendance 
                WHERE employee_id = %s AND date = %s
            """, (employee_id, date))
        else:
            # Insert or update attendance record with all fields
            cursor.execute("""
                INSERT INTO attendance (employee_id, date, status, comment, arrival_time, permission_type, start_date, end_date)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (employee_id, date)
                DO UPDATE SET 
                    status = EXCLUDED.status,
                    comment = EXCLUDED.comment,
                    arrival_time = EXCLUDED.arrival_time,
                    permission_type = EXCLUDED.permission_type,
                    start_date = EXCLUDED.start_date,
                    end_date = EXCLUDED.end_date
            """, (employee_id, date, status, comment, arrival_time, permission_type, start_date, end_date))
            
            # Check for accumulated delays if status is delay
            if status == 'delay':
                check_accumulated_delays(cursor, employee_id, date)
        
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()

@app.route('/api/attendance/<int:id>', methods=['DELETE', 'OPTIONS'])
def delete_attendance(id):
    if request.method == 'OPTIONS':
        return '', 204
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("DELETE FROM attendance WHERE id = %s", (id,))
        conn.commit()
        
        return jsonify({'success': True})
    finally:
        conn.close()

# ==================== API: REPORTS ====================

@app.route('/api/reports/archived-months', methods=['GET', 'OPTIONS'])
def reports_archived_months():
    if request.method == 'OPTIONS':
        return '', 204
    
    user_id = request.args.get('userId')
    if not user_id:
        return jsonify({'success': False, 'message': 'userId requerido'}), 400
        
    authorized_ids = get_authorized_user_ids(int(user_id))
    if not authorized_ids:
        return jsonify({'success': True, 'data': []})
        
    placeholders = ','.join(['%s' for _ in authorized_ids])
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(f"""
            SELECT 
                r.month, r.year, COUNT(*) as record_count
            FROM reports r
            JOIN attendance_roster ar ON r.employee_id = ar.employee_id
            WHERE ar.added_by_user_id IN ({placeholders})
            GROUP BY r.year, r.month
            ORDER BY r.year DESC, r.month DESC
        """, authorized_ids)
        
        data = []
        for row in cursor.fetchall():
            data.append({
                'month': row['month'], 
                'year': row['year'],
                'record_count': row['record_count']
            })
            
        return jsonify({'success': True, 'data': data})
    except Exception as e:
         print(f"Error en reports_archived_months: {e}")
         return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/reports', methods=['GET', 'POST', 'OPTIONS'])
def reports():
    if request.method == 'OPTIONS':
        return '', 204
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        if request.method == 'GET':
            user_id = request.args.get('userId')
            
            if not user_id:
                return jsonify({'success': False, 'message': 'userId requerido'}), 400
            
            authorized_ids = get_authorized_user_ids(int(user_id))
            
            if not authorized_ids:
                return jsonify({'success': False, 'message': 'No autorizado'}), 403
            
            month = request.args.get('month', 0)
            year = request.args.get('year', 2026)
            
            placeholders = ','.join(['%s' for _ in authorized_ids])
            
            # Obtener empleados del roster con sus datos de reportes
            # Tutores ven solo lo que ellos agregan
            # Admins ven solo lo que sus tutores supervisados agregan
            cursor.execute(f"""
                SELECT DISTINCT 
                    e.id, 
                    e.full_name, 
                    e.branch_id, 
                    b.name as branch_name,
                    r.data as report_data
                FROM attendance_roster ar
                JOIN employees e ON ar.employee_id = e.id
                LEFT JOIN branches b ON e.branch_id = b.id
                LEFT JOIN reports r ON r.employee_id = e.id 
                    AND r.month = %s AND r.year = %s
                WHERE ar.added_by_user_id IN ({placeholders})
                ORDER BY e.full_name ASC
            """, [month, year] + authorized_ids)
            
            employees = []
            for row in cursor:
                employees.append({
                    'id': row['id'],
                    'full_name': row['full_name'],
                    'branch_id': row['branch_id'],
                    'branch_name': row['branch_name'],
                    'report_data': row['report_data']  # Include report data
                })
            
            return jsonify({
                'success': True,
                'employees': employees
            })
        
        
        elif request.method == 'POST':
            data = request.json
            
            # Check if this is a full data update or incremental update
            if 'data' in data:
                # Full data update (old format)
                report_data = json.dumps(data.get('data'))
                employee_id = data.get('employee_id')
                month = data.get('month')
                year = data.get('year')
            else:
                # Incremental update (new format from frontend)
                employee_id = data.get('employee_id')
                month = data.get('month')
                year = data.get('year')
                update_type = data.get('type')  # 'faltantes', 'guias', 'tableros'
                key = data.get('key')  # day number or quincena/semana number
                status = data.get('status')
                comment = data.get('comment', '')
                
                # Fetch existing report data
                cursor.execute("""
                    SELECT data FROM reports 
                    WHERE employee_id = %s AND month = %s AND year = %s
                """, (employee_id, month, year))
                
                existing = cursor.fetchone()
                if existing and existing['data'] and existing['data'].strip():
                    try:
                        report_obj = json.loads(existing['data'])
                    except (json.JSONDecodeError, TypeError):
                        report_obj = {'faltantes': {}, 'guias': {}, 'tableros': {}}
                else:
                    report_obj = {'faltantes': {}, 'guias': {}, 'tableros': {}}
                
                # Ensure report_obj is a dict
                if not isinstance(report_obj, dict):
                    report_obj = {'faltantes': {}, 'guias': {}, 'tableros': {}}
                
                # Update the specific field
                if update_type not in report_obj:
                    report_obj[update_type] = {}
                
                if status == 'empty':
                    # Remove the entry
                    if str(key) in report_obj[update_type]:
                        del report_obj[update_type][str(key)]
                else:
                    report_obj[update_type][str(key)] = {
                        'status': status,
                        'comment': comment
                    }
                
                report_data = json.dumps(report_obj)
            
            cursor.execute("""
                INSERT INTO reports (employee_id, month, year, data)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (employee_id, month, year) 
                DO UPDATE SET data = EXCLUDED.data
            """, (employee_id, month, year, report_data))
            conn.commit()
            
            return jsonify({'success': True})
    finally:
        conn.close()

@app.route('/api/reports/data', methods=['GET', 'OPTIONS'])
def get_report_data():
    if request.method == 'OPTIONS':
        return '', 204
        
    employee_id = request.args.get('employeeId')
    month = request.args.get('month')
    year = request.args.get('year')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT data FROM reports
            WHERE employee_id = %s AND month = %s AND year = %s
        """, (employee_id, month, year))
        
        result = cursor.fetchone()
        
        if result and result['data']:
            data = json.loads(result['data'])
            return jsonify({'success': True, 'data': data})
        else:
            return jsonify({'success': True, 'data': None})
    finally:
        conn.close()

# ==================== API: ALERTS ====================

@app.route('/api/alerts', methods=['GET', 'OPTIONS'])
def get_alerts():
    if request.method == 'OPTIONS':
        return '', 204
        
    user_id = request.args.get('userId')
    mode = request.args.get('mode', 'unread') # unread, all
    if not user_id:
        return jsonify({'success': False, 'message': 'userId requerido'}), 400
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        if mode == 'all':
            # Get last 50 alerts regardless of read status
            cursor.execute("""
                SELECT id, details, is_read, created_at 
                FROM alerts 
                WHERE user_id = %s
                ORDER BY created_at DESC
                LIMIT 50
            """, (user_id,))
        else:
            # Get only unread
            cursor.execute("""
                SELECT id, details, is_read, created_at 
                FROM alerts 
                WHERE user_id = %s AND is_read = FALSE
                ORDER BY created_at DESC
            """, (user_id,))
        
        alerts = []
        for row in cursor:
            alerts.append({
                'id': row['id'],
                'details': json.loads(row['details']) if row['details'] else {},
                'is_read': row['is_read'],
                'created_at': row['created_at'].isoformat() if hasattr(row['created_at'], 'isoformat') else str(row['created_at'])
            })
            
        return jsonify({'success': True, 'data': alerts})
    finally:
        conn.close()

@app.route('/api/alerts/mark-read', methods=['POST', 'OPTIONS'])
def mark_alert_read():
    if request.method == 'OPTIONS':
        return '', 204
        
    data = request.json
    alert_id = data.get('alert_id')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("UPDATE alerts SET is_read = TRUE WHERE id = %s", (alert_id,))
        conn.commit()
        return jsonify({'success': True})
    finally:
        conn.close()

# ==================== API: REPORTS ====================

def get_report_context(year, month):
    """Obtiene días hábiles y feriados para un mes/año dado"""
    try:
        import urllib.request
        import json
        from datetime import date
        import calendar

        # Get holidays
        url = f"https://date.nager.at/api/v3/PublicHolidays/{year}/MX"
        req = urllib.request.Request(
            url, 
            headers={'User-Agent': 'Mozilla/5.0'}
        )
        holiday_dates = []
        try:
            with urllib.request.urlopen(req) as response:
                data = json.loads(response.read().decode())
                # Parse YYYY-MM-DD
                for h in data:
                    parts = h['date'].split('-')
                    holiday_dates.append(date(int(parts[0]), int(parts[1]), int(parts[2])))
        except:
            pass # Si falla API, asumimos 0 feriados

        num_days = calendar.monthrange(year, month)[1]
        business_days = []
        
        for d in range(1, num_days + 1):
            curr_date = date(year, month, d)
            # 0=Monday, 5=Saturday, 6=Sunday
            # Excluir fines de semana y feriados
            if curr_date.weekday() < 5 and curr_date not in holiday_dates:
                business_days.append(curr_date)
                
        return business_days, holiday_dates
    except Exception as e:
        print(f"Error ctx: {e}")
        return [], []

@app.route('/api/reports/tutors', methods=['GET', 'OPTIONS'])
def generate_tutor_report():
    """Generar reporte de resumen por tutor"""
    if request.method == 'OPTIONS':
        return '', 204
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        from flask import send_file
        import json
        
        user_id = request.args.get('userId')
        month = request.args.get('month') 
        year = request.args.get('year')
        tutor_id = request.args.get('tutorId')
        
        if not user_id or not month or not year:
            return jsonify({'success': False, 'message': 'Parámetros requeridos: userId, month, year'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT username, role FROM users WHERE id = %s", (user_id,))
        user_data = cursor.fetchone()
        
        if not user_data or user_data['role'] != 'admin':
            return jsonify({'success': False, 'message': 'No autorizado'}), 403
        
        allowed_users = ['helder mora', 'esthfania ramos']
        if user_data['username'].lower() not in allowed_users:
            return jsonify({'success': False, 'message': 'Solo Helder Mora y Esthfania Ramos pueden generar este reporte'}), 403
        
        target_month = int(month) + 1
        target_year = int(year)
        
        business_days, holidays = get_report_context(target_year, target_month)
        num_business_days = len(business_days)
        
        if tutor_id and tutor_id != 'all':
            cursor.execute("SELECT id, username FROM users WHERE id = %s AND supervisor_id = %s AND role = 'tutor_analista'", (tutor_id, user_id))
        else:
            cursor.execute("SELECT id, username FROM users WHERE supervisor_id = %s AND role = 'tutor_analista' ORDER BY username ASC", (user_id,))
        
        tutors = cursor.fetchall()
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Reporte Tutores {target_month}-{target_year}"
        
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        ws.merge_cells('A1:I1')
        ws['A1'] = f'REPORTE DE CUMPLIMIENTO POR TUTOR - {target_month}/{target_year}'
        ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = Alignment(horizontal='center')
        
        headers = ['Tutor', 'Colaboradores', 'Asistencias (Slots)', 'Incidencias', 'Reportes Completos', 
                   'Score Asistencia (50%)', 'Score Reportes (50%)', 'Cumplimiento %', 'Estado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        row = 4
        for tutor in tutors:
            tutor_id_val = tutor['id']
            tutor_name = tutor['username']
            
            cursor.execute("SELECT employee_id FROM attendance_roster WHERE added_by_user_id = %s", (tutor_id_val,))
            collabs = [r['employee_id'] for r in cursor.fetchall()]
            num_collabs = len(collabs)
            
            # --- 1. Asistencia Tutores (50%) ---
            total_attendance_slots_expected = num_business_days * num_collabs
            filled_slots = 0
            
            if collabs:
                placeholders = ','.join(['%s' for _ in collabs])
                cursor.execute(f"""
                    SELECT employee_id, date, status FROM attendance 
                    WHERE employee_id IN ({placeholders})
                    AND EXTRACT(MONTH FROM date) = %s AND EXTRACT(YEAR FROM date) = %s
                """, collabs + [target_month, target_year])
                
                att_map = {}
                for rec in cursor.fetchall():
                    att_map[(rec['employee_id'], rec['date'])] = rec['status']
                
                for cid in collabs:
                    for b_day in business_days:
                        # Any non-empty status counts
                        st = att_map.get((cid, b_day))
                        if st: # Not None/Empty
                            filled_slots += 1

            attendance_score = (filled_slots / total_attendance_slots_expected * 50) if total_attendance_slots_expected > 0 else 0
            
            # --- 2. Reportes Tutores (50%) ---
            complete_reports_count = 0
            
            if collabs:
                placeholders = ','.join(['%s' for _ in collabs])
                cursor.execute(f"""
                    SELECT employee_id, data as report_data FROM reports 
                    WHERE employee_id IN ({placeholders})
                    AND month = %s AND year = %s
                """, collabs + [target_month - 1, target_year]) # month 0-11
                
                report_records = {r['employee_id']: r['report_data'] for r in cursor.fetchall()}
                
                for cid in collabs:
                    rep_data_str = report_records.get(cid)
                    if rep_data_str:
                        try:
                            # Verify Completely Full (Any status for all requirements)
                            data = json.loads(rep_data_str)
                            is_full = True
                            
                            # Check Faltantes (All business days must have something)
                            faltantes = data.get('faltantes', {})
                            for b_day in business_days:
                                if not faltantes.get(str(b_day.day), {}).get('status'):
                                    is_full = False; break
                            
                            # Check Guias (2)
                            guias = data.get('guias', {})
                            for k in ['1', '2']:
                                if not guias.get(k, {}).get('status'):
                                    is_full = False; break
                                    
                            # Check Tableros (4)
                            tableros = data.get('tableros', {})
                            for k in ['1', '2', '3', '4']:
                                if not tableros.get(k, {}).get('status'):
                                    is_full = False; break
                                    
                            if is_full:
                                complete_reports_count += 1
                        except: pass
            
            report_score = (complete_reports_count / num_collabs * 50) if num_collabs > 0 else 0
            
            # --- Total ---
            cumplimiento = attendance_score + report_score
            
            cursor.execute("""
                SELECT COUNT(*) as count FROM incidents i
                JOIN attendance_roster ar ON i.reported_by = ar.employee_id
                WHERE ar.added_by_user_id = %s
                AND EXTRACT(MONTH FROM i.created_at) = %s AND EXTRACT(YEAR FROM i.created_at) = %s
            """, (tutor_id_val, target_month, target_year))
            tutor_incidencias = cursor.fetchone()['count']
            
            if cumplimiento >= 90: estado = "Excelente"; ecolor = "22C55E"
            elif cumplimiento >= 70: estado = "Bueno"; ecolor = "3B82F6"
            elif cumplimiento >= 50: estado = "Regular"; ecolor = "F59E0B"
            else: estado = "Bajo"; ecolor = "EF4444"
            
            vals = [tutor_name, num_collabs, filled_slots, tutor_incidencias, complete_reports_count,
                    f"{attendance_score:.1f}%", f"{report_score:.1f}%", f"{cumplimiento:.2f}%"]
            
            for c_idx, val in enumerate(vals, 1):
                ws.cell(row, c_idx, val).border = border
            
            ecell = ws.cell(row, 9, estado)
            ecell.border = border; ecell.fill = PatternFill(start_color=ecolor, end_color=ecolor, fill_type="solid")
            ecell.font = Font(color="FFFFFF", bold=True); ecell.alignment = Alignment(horizontal='center')
            
            row += 1
        
        ws.column_dimensions['A'].width = 25; ws.column_dimensions['B'].width = 15
        for c in ['C','D','E','F','G','H']: ws.column_dimensions[c].width = 18
        ws.column_dimensions['I'].width = 15
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        conn.close()
        
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'reporte_tutores_{target_month}_{target_year}.xlsx')
        
    except Exception as e:
        print(f"Error generating tutor report: {e}")
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/reports/collaborators', methods=['GET', 'OPTIONS'])
def generate_collaborator_report():
    """Generar reporte de resumen por colaborador"""
    if request.method == 'OPTIONS':
        return '', 204
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        from flask import send_file
        import json 
        
        user_id = request.args.get('userId')
        month = request.args.get('month') 
        year = request.args.get('year')
        collaborator_id = request.args.get('collaboratorId')
        
        if not user_id or not month or not year:
            return jsonify({'success': False, 'message': 'Parámetros requeridos: userId, month, year'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT username, role FROM users WHERE id = %s", (user_id,))
        user_data = cursor.fetchone()
        
        if not user_data or user_data['role'] != 'admin':
            return jsonify({'success': False, 'message': 'No autorizado'}), 403
        
        allowed_users = ['helder mora', 'esthfania ramos']
        if user_data['username'].lower() not in allowed_users:
            return jsonify({'success': False, 'message': 'Solo Helder Mora y Esthfania Ramos pueden generar este reporte'}), 403
        
        target_month = int(month) + 1
        target_year = int(year)
        
        # Contexto (Días hábiles)
        business_days, holidays = get_report_context(target_year, target_month)
        num_business_days = len(business_days)
        
        # Obtener tutores
        cursor.execute("SELECT id FROM users WHERE supervisor_id = %s AND role = 'tutor_analista'", (user_id,))
        tutor_ids = [row['id'] for row in cursor.fetchall()]
        
        if not tutor_ids:
            return jsonify({'success': False, 'message': 'No se encontraron tutores supervisados'}), 404
        
        placeholders = ','.join(['%s' for _ in tutor_ids])
        
        sql_collabs = f"SELECT DISTINCT e.id, e.full_name, b.name as branch_name FROM employees e LEFT JOIN branches b ON e.branch_id = b.id JOIN attendance_roster ar ON e.id = ar.employee_id WHERE ar.added_by_user_id IN ({placeholders})"
        params = list(tutor_ids)
        
        if collaborator_id and collaborator_id != 'all':
            sql_collabs += " AND e.id = %s"
            params.append(collaborator_id)
            
        sql_collabs += " ORDER BY e.full_name ASC"
        cursor.execute(sql_collabs, params)
        collaborators = cursor.fetchall()
        
        if not collaborators:
            return jsonify({'success': False, 'message': 'No se encontraron colaboradores'}), 404
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"Reporte Colaboradores {target_month}-{target_year}"
        
        header_fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        ws.merge_cells('A1:K1')
        ws['A1'] = f'REPORTE DE CUMPLIMIENTO POR COLABORADOR - {target_month}/{target_year}'
        ws['A1'].font = Font(bold=True, size=14); ws['A1'].alignment = Alignment(horizontal='center')
        
        headers = ['Colaborador', 'Sucursal', 'Asistencias', 'Faltas', 'Vacaciones', 'Permisos', 'Incapacidades', 'Incidencias', 'Reportes (pts)', 'Cumplimiento %', 'Estado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
        row = 4
        for collab in collaborators:
            cid = collab['id']
            cname = collab['full_name']
            bname = collab['branch_name'] or 'Sin sucursal'
            
            # --- 1. Asistencia (50%) ---
            cursor.execute("SELECT date, status FROM attendance WHERE employee_id = %s AND EXTRACT(MONTH FROM date) = %s AND EXTRACT(YEAR FROM date) = %s", (cid, target_month, target_year))
            att_records = {r['date']: r['status'] for r in cursor.fetchall()}
            
            valid_attendance_count = 0
            c_asist = 0; c_faltas = 0; c_vac = 0; c_perm = 0; c_incap = 0
            
            for b_day in business_days:
                st = att_records.get(b_day)
                if st == 'present' or st == 'delay':
                    valid_attendance_count += 1; c_asist += 1
                elif st == 'vacation':
                    valid_attendance_count += 1; c_vac += 1
                elif st == 'permission':
                    valid_attendance_count += 1; c_perm += 1
                elif st == 'absent':
                    c_faltas += 1
                elif st == 'incapacity':
                    c_incap += 1
            
            attendance_score = (valid_attendance_count / num_business_days * 50) if num_business_days > 0 else 0
            
            # --- 2. Reportes (50%) ---
            cursor.execute("SELECT data as report_data FROM reports WHERE employee_id=%s AND month=%s AND year=%s", (cid, target_month-1, target_year))
            rep_row = cursor.fetchone()
            
            valid_report_items = 0
            expected_report_items = num_business_days + 2 + 4 # Dias + 2 Guias + 4 Tableros
            
            if rep_row and rep_row['report_data']:
                try:
                    data = json.loads(rep_row['report_data'])
                    
                    # Faltantes (Dias habiles)
                    faltantes = data.get('faltantes', {})
                    for b_day in business_days:
                         if faltantes.get(str(b_day.day), {}).get('status') == 'check':
                             valid_report_items += 1
                    
                    # Guias (2 envios) - Keys '1', '2'
                    guias = data.get('guias', {})
                    for k in ['1', '2']:
                         if guias.get(k, {}).get('status') == 'check':
                             valid_report_items += 1

                    # Tableros (4 evidencias) - Keys '1', '2', '3', '4'
                    tableros = data.get('tableros', {})
                    for k in ['1', '2', '3', '4']:
                         if tableros.get(k, {}).get('status') == 'check':
                             valid_report_items += 1
                             
                except: pass
            
            report_score = (valid_report_items / expected_report_items * 50) if expected_report_items > 0 else 0
            
            # --- Total ---
            cumplimiento = attendance_score + report_score
            
            # Incidencias (Solo contar)
            cursor.execute("SELECT COUNT(*) as count FROM incidents WHERE reported_by = %s AND EXTRACT(MONTH FROM created_at) = %s AND EXTRACT(YEAR FROM created_at) = %s", (cid, target_month, target_year))
            c_incidencias = cursor.fetchone()['count']
            
            if cumplimiento >= 90: estado = "Excelente"; ecolor = "22C55E"
            elif cumplimiento >= 70: estado = "Bueno"; ecolor = "3B82F6"
            elif cumplimiento >= 50: estado = "Regular"; ecolor = "F59E0B"
            else: estado = "Bajo"; ecolor = "EF4444"
            
            ws.cell(row, 1, cname).border = border
            ws.cell(row, 2, bname).border = border
            ws.cell(row, 3, c_asist).border = border
            ws.cell(row, 4, c_faltas).border = border
            ws.cell(row, 5, c_vac).border = border
            ws.cell(row, 6, c_perm).border = border
            ws.cell(row, 7, c_incap).border = border
            ws.cell(row, 8, c_incidencias).border = border
            ws.cell(row, 9, valid_report_items).border = border
            ws.cell(row, 10, f"{cumplimiento:.2f}%").border = border
            
            ecell = ws.cell(row, 11, estado)
            ecell.border = border; ecell.fill = PatternFill(start_color=ecolor, end_color=ecolor, fill_type="solid")
            ecell.font = Font(color="FFFFFF", bold=True); ecell.alignment = Alignment(horizontal='center')
            
            row += 1
            
        ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 20
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        conn.close()
        
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'reporte_colaboradores_{target_month}_{target_year}.xlsx')
        
    except Exception as e:
        print(f"Error generating collaborator report: {e}")
        import traceback; traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/reports/available-months', methods=['GET', 'OPTIONS'])
def get_available_report_months():
    """Obtener meses disponibles para reportes"""
    if request.method == 'OPTIONS':
        return '', 204
    
    try:
        user_id = request.args.get('userId')
        
        if not user_id:
            return jsonify({'success': False, 'message': 'userId requerido'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verificar que el usuario es administrador
        cursor.execute("SELECT role FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        
        if not user or user['role'] != 'admin':
            return jsonify({'success': False, 'message': 'No autorizado'}), 403
        
        # Obtener meses con datos de asistencia
        cursor.execute("""
            SELECT DISTINCT 
                EXTRACT(MONTH FROM date) as month,
                EXTRACT(YEAR FROM date) as year
            FROM attendance
            ORDER BY year DESC, month DESC
            LIMIT 12
        """)
        
        months = []
        for row in cursor.fetchall():
            months.append({
                'month': int(row['month']) - 1,  # Convertir a formato JS (0-11)
                'year': int(row['year'])
            })
        
        conn.close()
        
        return jsonify({'success': True, 'data': months})
        
    except Exception as e:
        print(f"Error getting available months: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# ==================== CORS Headers ====================

@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response

# ==================== REPORTES ESPECÍFICOS ====================

@app.route('/api/reports/absences', methods=['GET', 'OPTIONS'])
def generate_absences_report():
    """Generar reporte de faltas por colaborador"""
    if request.method == 'OPTIONS':
        return '', 204
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        from flask import send_file
        
        user_id = request.args.get('userId')
        month = request.args.get('month')  # 0-11
        year = request.args.get('year')
        collaborator_id = request.args.get('collaboratorId')  # Optional
        
        if not user_id or not month or not year:
            return jsonify({'success': False, 'message': 'Parámetros requeridos: userId, month, year'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verificar que el usuario es administrador
        cursor.execute("SELECT username, role FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        
        if not user or user['role'] != 'admin':
            return jsonify({'success': False, 'message': 'No autorizado'}), 403
        
        # Verificar que es Helder Mora o Esthfania Ramos (case-insensitive)
        allowed_users = ['helder mora', 'esthfania ramos']
        if user['username'].lower() not in allowed_users:
            return jsonify({'success': False, 'message': 'Solo Helder Mora y Esthfania Ramos pueden generar este reporte'}), 403
        
        # Convertir mes de JS (0-11) a SQL (1-12)
        target_month = int(month) + 1
        target_year = int(year)
        
        # Query para obtener faltas
        if collaborator_id and collaborator_id != 'all':
            cursor.execute("""
                SELECT 
                    e.full_name,
                    b.name as branch_name,
                    COUNT(*) as total_faltas,
                    STRING_AGG(TO_CHAR(a.date, 'DD/MM/YYYY'), ', ') as fechas
                FROM attendance a
                JOIN employees e ON a.employee_id = e.id
                LEFT JOIN branches b ON e.branch_id = b.id
                WHERE a.status = 'absent'
                    AND EXTRACT(MONTH FROM a.date) = %s
                    AND EXTRACT(YEAR FROM a.date) = %s
                    AND e.id = %s
                GROUP BY e.id, e.full_name, b.name
                ORDER BY e.full_name
            """, (target_month, target_year, collaborator_id))
        else:
            cursor.execute("""
                SELECT 
                    e.full_name,
                    b.name as branch_name,
                    COUNT(*) as total_faltas,
                    STRING_AGG(TO_CHAR(a.date, 'DD/MM/YYYY'), ', ') as fechas
                FROM attendance a
                JOIN employees e ON a.employee_id = e.id
                LEFT JOIN branches b ON e.branch_id = b.id
                WHERE a.status = 'absent'
                    AND EXTRACT(MONTH FROM a.date) = %s
                    AND EXTRACT(YEAR FROM a.date) = %s
                GROUP BY e.id, e.full_name, b.name
                ORDER BY e.full_name
            """, (target_month, target_year))
        
        absences = cursor.fetchall()
        
        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Faltas"
        
        # Estilos
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = ['Colaborador', 'Sucursal', 'Total Faltas', 'Fechas']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos
        for row_idx, absence in enumerate(absences, 2):
            ws.cell(row=row_idx, column=1, value=absence['full_name']).border = border
            ws.cell(row=row_idx, column=2, value=absence['branch_name'] or 'N/A').border = border
            ws.cell(row=row_idx, column=3, value=absence['total_faltas']).border = border
            ws.cell(row=row_idx, column=4, value=absence['fechas']).border = border
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 50
        
        # Guardar
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        conn.close()
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'reporte_faltas_{target_month}_{target_year}.xlsx'
        )
        
    except Exception as e:
        print(f"Error generating absences report: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/reports/vacations', methods=['GET', 'OPTIONS'])
def generate_vacations_report():
    """Generar reporte de vacaciones por colaborador"""
    if request.method == 'OPTIONS':
        return '', 204
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        from flask import send_file
        
        user_id = request.args.get('userId')
        month = request.args.get('month')  # 0-11
        year = request.args.get('year')
        collaborator_id = request.args.get('collaboratorId')  # Optional
        
        if not user_id or not month or not year:
            return jsonify({'success': False, 'message': 'Parámetros requeridos: userId, month, year'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verificar que el usuario es administrador
        cursor.execute("SELECT username, role FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        
        if not user or user['role'] != 'admin':
            return jsonify({'success': False, 'message': 'No autorizado'}), 403
        
        # Verificar que es Helder Mora o Esthfania Ramos (case-insensitive)
        allowed_users = ['helder mora', 'esthfania ramos']
        if user['username'].lower() not in allowed_users:
            return jsonify({'success': False, 'message': 'Solo Helder Mora y Esthfania Ramos pueden generar este reporte'}), 403
        
        # Convertir mes de JS (0-11) a SQL (1-12)
        target_month = int(month) + 1
        target_year = int(year)
        
        # Query para obtener vacaciones
        if collaborator_id and collaborator_id != 'all':
            cursor.execute("""
                SELECT 
                    e.full_name,
                    b.name as branch_name,
                    COUNT(*) as total_dias_vacaciones,
                    STRING_AGG(TO_CHAR(a.date, 'DD/MM/YYYY'), ', ') as fechas
                FROM attendance a
                JOIN employees e ON a.employee_id = e.id
                LEFT JOIN branches b ON e.branch_id = b.id
                WHERE a.status = 'vacation'
                    AND EXTRACT(MONTH FROM a.date) = %s
                    AND EXTRACT(YEAR FROM a.date) = %s
                    AND e.id = %s
                GROUP BY e.id, e.full_name, b.name
                ORDER BY e.full_name
            """, (target_month, target_year, collaborator_id))
        else:
            cursor.execute("""
                SELECT 
                    e.full_name,
                    b.name as branch_name,
                    COUNT(*) as total_dias_vacaciones,
                    STRING_AGG(TO_CHAR(a.date, 'DD/MM/YYYY'), ', ') as fechas
                FROM attendance a
                JOIN employees e ON a.employee_id = e.id
                LEFT JOIN branches b ON e.branch_id = b.id
                WHERE a.status = 'vacation'
                    AND EXTRACT(MONTH FROM a.date) = %s
                    AND EXTRACT(YEAR FROM a.date) = %s
                GROUP BY e.id, e.full_name, b.name
                ORDER BY e.full_name
            """, (target_month, target_year))
        
        vacations = cursor.fetchall()
        
        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Vacaciones"
        
        # Estilos
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = ['Colaborador', 'Sucursal', 'Total Días', 'Fechas']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos
        for row_idx, vacation in enumerate(vacations, 2):
            ws.cell(row=row_idx, column=1, value=vacation['full_name']).border = border
            ws.cell(row=row_idx, column=2, value=vacation['branch_name'] or 'N/A').border = border
            ws.cell(row=row_idx, column=3, value=vacation['total_dias_vacaciones']).border = border
            ws.cell(row=row_idx, column=4, value=vacation['fechas']).border = border
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 50
        
        # Guardar
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        conn.close()
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'reporte_vacaciones_{target_month}_{target_year}.xlsx'
        )
        
    except Exception as e:
        print(f"Error generating vacations report: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/reports/incidents', methods=['GET', 'OPTIONS'])
def generate_incidents_report():
    """Generar reporte de incidencias por sucursal"""
    if request.method == 'OPTIONS':
        return '', 204
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        from flask import send_file
        
        user_id = request.args.get('userId')
        month = request.args.get('month')  # 0-11
        year = request.args.get('year')
        branch_id = request.args.get('branchId')  # Optional
        
        if not user_id or not month or not year:
            return jsonify({'success': False, 'message': 'Parámetros requeridos: userId, month, year'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verificar que el usuario es administrador
        cursor.execute("SELECT username, role FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()
        
        if not user or user['role'] != 'admin':
            return jsonify({'success': False, 'message': 'No autorizado'}), 403
        
        # Verificar que es Helder Mora o Esthfania Ramos (case-insensitive)
        allowed_users = ['helder mora', 'esthfania ramos']
        if user['username'].lower() not in allowed_users:
            return jsonify({'success': False, 'message': 'Solo Helder Mora y Esthfania Ramos pueden generar este reporte'}), 403
        
        # Convertir mes de JS (0-11) a SQL (1-12)
        target_month = int(month) + 1
        target_year = int(year)
        
        # Query para obtener incidencias (las incidencias son por sucursal, no por colaborador)
        if branch_id and branch_id != 'all':
            cursor.execute("""
                SELECT 
                    i.id,
                    b.name as sucursal,
                    i.type as tipo,
                    i.description as descripcion,
                    i.status as estatus,
                    TO_CHAR(i.created_at, 'DD/MM/YYYY HH24:MI') as fecha_registro,
                    u.username as reportado_por
                FROM incidents i
                LEFT JOIN branches b ON i.branch_id = b.id
                LEFT JOIN users u ON i.reported_by = u.id
                WHERE EXTRACT(MONTH FROM i.created_at) = %s
                    AND EXTRACT(YEAR FROM i.created_at) = %s
                    AND i.branch_id = %s
                ORDER BY i.created_at DESC
            """, (target_month, target_year, branch_id))
        else:
            cursor.execute("""
                SELECT 
                    i.id,
                    b.name as sucursal,
                    i.type as tipo,
                    i.description as descripcion,
                    i.status as estatus,
                    TO_CHAR(i.created_at, 'DD/MM/YYYY HH24:MI') as fecha_registro,
                    u.username as reportado_por
                FROM incidents i
                LEFT JOIN branches b ON i.branch_id = b.id
                LEFT JOIN users u ON i.reported_by = u.id
                WHERE EXTRACT(MONTH FROM i.created_at) = %s
                    AND EXTRACT(YEAR FROM i.created_at) = %s
                ORDER BY i.created_at DESC
            """, (target_month, target_year))
        
        incidents = cursor.fetchall()
        
        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Incidencias"
        
        # Estilos
        header_fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados (sin columna de colaborador)
        headers = ['ID', 'Sucursal', 'Tipo', 'Descripción', 'Estatus', 'Fecha Registro', 'Reportado Por']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Datos
        for row_idx, incident in enumerate(incidents, 2):
            ws.cell(row=row_idx, column=1, value=incident['id']).border = border
            ws.cell(row=row_idx, column=2, value=incident['sucursal'] or 'N/A').border = border
            ws.cell(row=row_idx, column=3, value=incident['tipo']).border = border
            ws.cell(row=row_idx, column=4, value=incident['descripcion']).border = border
            
            # Color según estatus
            status_cell = ws.cell(row=row_idx, column=5, value=incident['estatus'])
            status_cell.border = border
            if incident['estatus'] == 'activa':
                status_cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            elif incident['estatus'] == 'resuelta':
                status_cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
            
            ws.cell(row=row_idx, column=6, value=incident['fecha_registro']).border = border
            ws.cell(row=row_idx, column=7, value=incident['reportado_por'] or 'N/A').border = border
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 20
        
        # Guardar
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        conn.close()
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'reporte_incidencias_{target_month}_{target_year}.xlsx'
        )
        
    except Exception as e:
        print(f"Error generating incidents report: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/branches', methods=['GET', 'OPTIONS'])
def fetch_branches_list():
    """Obtener lista de sucursales"""
    if request.method == 'OPTIONS':
        return '', 204
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Obtener sucursales de la tabla branches
        cursor.execute("""
            SELECT id, name, location, created_at
            FROM branches
            ORDER BY name
        """)
        
        branches = []
        for row in cursor.fetchall():
            branches.append({
                'id': row['id'],
                'name': row['name'],
                'location': row['location'] if 'location' in row.keys() else None
            })
        
        conn.close()
        
        return jsonify({'success': True, 'data': branches})
        
    except Exception as e:
        print(f"Error fetching branches: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

# ==================== CAMBIOS EN SUCURSALES (ROTACIÓN) ====================

@app.route('/api/branch-changes', methods=['GET', 'OPTIONS'])
def get_branch_changes():
    """Obtener cambios pendientes en sucursales"""
    if request.method == 'OPTIONS':
        return '', 204
    
    try:
        user_id = request.args.get('userId')
        
        if not user_id:
            return jsonify({'success': False, 'message': 'userId requerido'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Obtener cambios no procesados
        cursor.execute("""
            SELECT 
                id,
                branch_id,
                branch_name,
                change_type,
                employee_name,
                TO_CHAR(hire_date, 'DD/MM/YYYY') as hire_date,
                description,
                TO_CHAR(created_at, 'DD/MM/YYYY HH24:MI') as created_at
            FROM branch_changes
            WHERE processed = FALSE
            ORDER BY created_at DESC
        """)
        
        changes = []
        for row in cursor.fetchall():
            changes.append({
                'id': row['id'],
                'branch_id': row['branch_id'],
                'branch_name': row['branch_name'],
                'change_type': row['change_type'],
                'employee_name': row['employee_name'],
                'hire_date': row['hire_date'],
                'description': row['description'],
                'created_at': row['created_at']
            })
        
        conn.close()
        
        return jsonify({'success': True, 'data': changes})
        
    except Exception as e:
        print(f"Error fetching branch changes: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/branch-changes/<int:change_id>', methods=['DELETE', 'OPTIONS'])
def process_branch_change(change_id):
    """Marcar un cambio como procesado (eliminarlo de la lista)"""
    if request.method == 'OPTIONS':
        return '', 204
    
    try:
        data = request.get_json()
        user_id = data.get('userId')
        
        if not user_id:
            return jsonify({'success': False, 'message': 'userId requerido'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Marcar como procesado
        cursor.execute("""
            UPDATE branch_changes
            SET processed = TRUE,
                processed_at = CURRENT_TIMESTAMP,
                processed_by = %s
            WHERE id = %s
        """, (user_id, change_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Cambio procesado correctamente'})
        
    except Exception as e:
        print(f"Error processing branch change: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ==================== MAIN ====================

# ==================== DASHBOARD STATS ====================

@app.route('/api/dashboard/stats', methods=['GET', 'OPTIONS'])
def get_dashboard_stats():
    if request.method == 'OPTIONS':
        return '', 204
    
    user_id = request.args.get('userId')
    if not user_id:
        return jsonify({'success': False, 'message': 'userId requerido'}), 400
        
    authorized_ids = get_authorized_user_ids(int(user_id))
    if not authorized_ids:
        return jsonify({'success': True, 'data': {
            'asistencias': 0, 'faltas': 0, 'vacaciones': 0, 'permisos': 0, 'incapacidades': 0, 'incidencias_activas': 0, 'retardos': 0
        }})
        
    placeholders = ','.join(['%s' for _ in authorized_ids])
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        
        # Stats de asistencia de HOY
        sql_base = f"""
            SELECT a.status, COUNT(*) as count
            FROM attendance a
            JOIN attendance_roster ar ON a.employee_id = ar.employee_id
            WHERE ar.added_by_user_id IN ({placeholders})
            AND a.date = %s
            GROUP BY a.status
        """
        
        cursor.execute(sql_base, authorized_ids + [today])
        
        counts = {row['status']: row['count'] for row in cursor.fetchall()}
        
        # Incidencias activas (TOTAL, no solo de hoy)
        cursor.execute(f"""
            SELECT COUNT(*) as count 
            FROM incidents i
            WHERE (i.reported_by IN ({placeholders}) OR i.branch_id IN (
                SELECT e.branch_id FROM employees e 
                JOIN attendance_roster ar ON e.id = ar.employee_id 
                WHERE ar.added_by_user_id IN ({placeholders})
            ))
            AND i.status = 'activa'
        """, authorized_ids * 2) 
        
        incidencias = cursor.fetchone()['count']
        
        return jsonify({'success': True, 'data': {
            'asistencias': counts.get('present', 0),
            'faltas': counts.get('absent', 0),
            'vacaciones': counts.get('vacation', 0),
            'permisos': counts.get('permission', 0),
            'incapacidades': counts.get('incapacity', 0),
            'retardos': counts.get('delay', 0),
            'incidencias_activas': incidencias
        }})
        
    except Exception as e:
        print(f"Error getting dashboard stats: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/dashboard/delays', methods=['GET', 'OPTIONS'])
def get_dashboard_delays():
    if request.method == 'OPTIONS':
        return '', 204
        
    user_id = request.args.get('userId')
    if not user_id:
        return jsonify({'success': False, 'message': 'userId requerido'}), 400
        
    authorized_ids = get_authorized_user_ids(int(user_id))
    if not authorized_ids:
         return jsonify({'success': True, 'data': []})

    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        today = datetime.now().strftime('%Y-%m-%d')
        placeholders = ','.join(['%s' for _ in authorized_ids])
        
        cursor.execute(f"""
            SELECT 
                e.full_name,
                b.name as branch_name,
                a.arrival_time,
                a.comment
            FROM attendance a
            JOIN employees e ON a.employee_id = e.id
            LEFT JOIN branches b ON e.branch_id = b.id
            JOIN attendance_roster ar ON e.id = ar.employee_id
            WHERE ar.added_by_user_id IN ({placeholders})
            AND a.date = %s
            AND a.status = 'delay'
            ORDER BY e.full_name ASC
        """, authorized_ids + [today])
        
        delays = []
        for row in cursor.fetchall():
            delays.append({
                'full_name': row['full_name'],
                'branch_name': row['branch_name'],
                'arrival_time': row['arrival_time'],
                'comment': row['comment']
            })
            
        return jsonify({'success': True, 'data': delays})
    except Exception as e:
        print(f"Error getting delays: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/holidays/<int:year>', methods=['GET', 'OPTIONS'])
def get_holidays(year):
    if request.method == 'OPTIONS':
        return '', 204
        
    try:
        import urllib.request
        
        # URL de la API de días festivos
        url = f"https://date.nager.at/api/v3/PublicHolidays/{year}/MX"
        
        # Configurar request con User-Agent para evitar bloqueo 403
        req = urllib.request.Request(
            url, 
            data=None, 
            headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
        )
        
        with urllib.request.urlopen(req) as response:
            data = json.loads(response.read().decode())
            # Simplificar datos: solo devolver fechas
            holidays = [item['date'] for item in data] 
            return jsonify({'success': True, 'data': holidays})
            
    except Exception as e:
        print(f"Error fetching holidays: {e}")
        # En caso de error, devolver lista vacía para no romper el frontend
        return jsonify({'success': True, 'data': []})

if __name__ == '__main__':
    port = int(os.getenv('PORT', 3000))
    app.run(host='0.0.0.0', port=port, debug=True)
